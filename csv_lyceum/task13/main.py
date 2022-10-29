import csv
import openpyxl


data = openpyxl.load_workbook("data.xlsx", data_only=True)

sheet_obj = data.active
max_row, max_column = sheet_obj.max_row, sheet_obj.max_column


with open("output.csv", "w", encoding="utf-8") as file:
    writer = csv.writer(file, delimiter=';', quotechar='"')

    for r in range(1, max_row + 1):

        headers = []
        for c in range(1, max_column + 1):
            headers.append(sheet_obj.cell(row=r, column=c).value)

        writer.writerow(headers)
